"""
export_to_excel.py
------------------
Fetches today's avalanche forecasts from all active NAC zones and exports
to a formatted Excel workbook.

Sheets:
  1. Forecast Summary  — one row per zone; danger by elevation band
  2. Avalanche Problems — one row per problem; type, aspects, elevations, likelihood, size
  3. Stats             — center counts, danger breakdown

Usage:
    python export_to_excel.py [output.xlsx]
"""

from __future__ import annotations

import logging
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from forecast_parser import ForecastClient, DangerLevel

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", start_color="2E4057")
DATA_FONT    = Font(name="Arial", size=10)
SMALL_FONT   = Font(name="Arial", size=9)
GROUP_FONT   = Font(name="Arial", bold=True, size=11, color="FFFFFF")
GROUP_FILL   = PatternFill("solid", start_color="1A3A5C")

_THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Danger level → Excel bg/fg hex (no leading #)
DANGER_STYLE: dict[str, dict] = {
    "no_rating":    {"bg": "DDDDDD", "fg": "333333"},
    "low":          {"bg": "78B943", "fg": "000000"},
    "moderate":     {"bg": "F4C025", "fg": "000000"},
    "considerable": {"bg": "F58220", "fg": "000000"},
    "high":         {"bg": "E3001F", "fg": "FFFFFF"},
    "extreme":      {"bg": "1A1A1A", "fg": "FFFFFF"},
}

LIKELIHOOD_ORDER = [
    "almost_certain", "very_likely", "likely", "possible", "unlikely", ""
]


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _header(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = HEADER_FONT
    c.fill  = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _cell(ws, row, col, value, bold=False, align="left", wrap=False, size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = Font(name="Arial", size=size, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = BORDER
    return c


def _danger_cell(ws, row, col, label: str):
    style   = DANGER_STYLE.get(label, DANGER_STYLE["no_rating"])
    display = label.replace("_", " ").title() if label != "no_rating" else "No Rating"
    c = ws.cell(row=row, column=col, value=display)
    c.fill  = PatternFill("solid", start_color=style["bg"])
    c.font  = Font(name="Arial", size=10, bold=True, color=style["fg"])
    c.alignment = Alignment(horizontal="center", vertical="top")
    c.border = BORDER
    return c


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 1 — Forecast Summary
# ---------------------------------------------------------------------------

SUMMARY_HEADERS = [
    "Center ID", "Center Name", "State", "Zone",
    "Valid Date", "Issued (UTC)",
    "Alpine", "Treeline", "Below Treeline", "Max Danger",
    "# Problems", "Bottom Line",
]
SUMMARY_WIDTHS = [10, 32, 7, 30, 12, 18, 16, 16, 16, 16, 10, 60]


def build_summary(wb, forecasts: list):
    ws = wb.active
    ws.title = "Forecast Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    for ci, h in enumerate(SUMMARY_HEADERS, 1):
        _header(ws, 1, ci, h)

    # Sort by center, then zone
    for row, fc in enumerate(forecasts, 2):
        d = fc.to_dict()
        dr = {r["elevation_band"]: r for r in d["danger_ratings"]}

        col = 1
        _cell(ws, row, col, d["center_id"],   align="center"); col += 1
        _cell(ws, row, col, d["center_name"]);                 col += 1

        # State from zone list (not in to_dict, pull from forecast_zone if available)
        state = ""
        _cell(ws, row, col, state, align="center");            col += 1
        _cell(ws, row, col, d["zone_name"]);                   col += 1
        _cell(ws, row, col, d["valid_for_date"][:10], align="center"); col += 1
        _cell(ws, row, col, d["issued_at"][:16].replace("T", " "), align="center"); col += 1

        for band in ("alpine", "treeline", "below_treeline"):
            label = dr.get(band, {}).get("danger_label", "no_rating")
            _danger_cell(ws, row, col, label); col += 1

        _danger_cell(ws, row, col, d["max_danger"]); col += 1
        _cell(ws, row, col, len(d["avalanche_problems"]), align="center"); col += 1
        _cell(ws, row, col, d["bottom_line"] or "", wrap=True, size=9); col += 1

        ws.row_dimensions[row].height = 60

    ws.freeze_panes = ws.cell(row=2, column=1)
    _set_col_widths(ws, SUMMARY_WIDTHS)


# ---------------------------------------------------------------------------
# Sheet 2 — Avalanche Problems
# ---------------------------------------------------------------------------

PROB_HEADERS = [
    "Center ID", "Zone",
    "Problem #", "Problem Type",
    "Likelihood", "Size Min (D)", "Size Max (D)",
    "Aspects", "Elevation Bands",
    "Discussion",
]
PROB_WIDTHS = [10, 30, 10, 20, 16, 12, 12, 30, 24, 70]


def build_problems(wb, forecasts: list):
    ws = wb.create_sheet("Avalanche Problems")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    for ci, h in enumerate(PROB_HEADERS, 1):
        _header(ws, 1, ci, h)

    row = 2
    for fc in forecasts:
        d = fc.to_dict()

        if not d["avalanche_problems"]:
            _cell(ws, row, 1, d["center_id"], align="center")
            _cell(ws, row, 2, d["zone_name"])
            _cell(ws, row, 3, "—", align="center")
            _cell(ws, row, 4, "No problems listed", italic=True)
            ws.row_dimensions[row].height = 18
            row += 1
            continue

        for prob in d["avalanche_problems"]:
            col = 1
            _cell(ws, row, col, d["center_id"], align="center");       col += 1
            _cell(ws, row, col, d["zone_name"]);                       col += 1
            _cell(ws, row, col, prob["rank"], align="center");         col += 1
            _cell(ws, row, col, prob["problem_type"], bold=True);      col += 1

            # Likelihood cell — color coded
            lik = (prob["likelihood"] or "").replace("_", " ").title()
            c = _cell(ws, row, col, lik, align="center")
            lik_colors = {
                "Almost Certain": ("E3001F", "FFFFFF"),
                "Very Likely":    ("F58220", "000000"),
                "Likely":         ("F4C025", "000000"),
                "Possible":       ("78B943", "000000"),
                "Unlikely":       ("CCCCCC", "000000"),
            }
            if lik in lik_colors:
                bg, fg = lik_colors[lik]
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name="Arial", size=10, bold=True, color=fg)
            col += 1

            _cell(ws, row, col, prob["size_min"], align="center");     col += 1
            _cell(ws, row, col, prob["size_max"], align="center");     col += 1
            _cell(ws, row, col, ", ".join(prob["aspects"] or []), align="center"); col += 1
            _cell(ws, row, col, ", ".join(
                b.replace("_", " ").title() for b in (prob["elevation_bands"] or [])
            ));                                                        col += 1
            _cell(ws, row, col, prob["discussion"] or "", wrap=True, size=9); col += 1

            ws.row_dimensions[row].height = 60
            row += 1

    ws.freeze_panes = ws.cell(row=2, column=1)
    _set_col_widths(ws, PROB_WIDTHS)


# ---------------------------------------------------------------------------
# Sheet 3 — Stats
# ---------------------------------------------------------------------------

def build_stats(wb, forecasts: list):
    ws = wb.create_sheet("Stats")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    def row_pair(r, label, value, bold=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", size=10, bold=bold)
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = Font(name="Arial", size=10, bold=bold)
        c2.alignment = Alignment(horizontal="center")
        c2.border = BORDER

    _header(ws, 1, 1, "Metric")
    _header(ws, 1, 2, "Value")
    ws.row_dimensions[1].height = 28

    # Center count
    centers = {fc.center_id for fc in forecasts}
    total_problems = sum(len(fc.avalanche_problems) for fc in forecasts)

    row_pair(2, "Fetch Date", date.today().isoformat())
    row_pair(3, "Total Zones", len(forecasts), bold=True)
    row_pair(4, "Unique Centers", len(centers))
    row_pair(5, "Total Avalanche Problems", total_problems)

    # Danger breakdown
    ws.cell(row=7, column=1).value = "Danger Breakdown"
    ws.cell(row=7, column=1).font = Font(name="Arial", size=10, bold=True)

    danger_counts: dict[str, int] = defaultdict(int)
    for fc in forecasts:
        danger_counts[fc.max_danger().label] += 1

    r = 8
    for label, style in DANGER_STYLE.items():
        if label == "no_rating":
            continue
        cnt = danger_counts.get(label, 0)
        c1 = ws.cell(row=r, column=1, value=label.replace("_", " ").title())
        c1.fill = PatternFill("solid", start_color=style["bg"])
        c1.font = Font(name="Arial", size=10, bold=True, color=style["fg"])
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=cnt)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(horizontal="center")
        c2.border = BORDER
        r += 1

    nr = danger_counts.get("no_rating", 0)
    row_pair(r, "No Rating", nr)
    r += 2

    # Problem type breakdown
    ws.cell(row=r, column=1).value = "Problems by Type"
    ws.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True)
    r += 1
    type_counts: dict[str, int] = defaultdict(int)
    for fc in forecasts:
        for prob in fc.avalanche_problems:
            type_counts[prob.problem_type] += 1
    for ptype, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        row_pair(r, ptype, cnt)
        r += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    output_path = sys.argv[1] if len(sys.argv) > 1 else \
        f"avi_forecasts_{date.today()}.xlsx"

    client = ForecastClient()
    logger.info("Fetching all active zone forecasts…")
    forecasts = client.fetch_all()

    if not forecasts:
        logger.error("No forecasts fetched.")
        sys.exit(1)

    logger.info("Fetched %d zone forecasts across %d centers.",
                len(forecasts), len({f.center_id for f in forecasts}))

    wb = openpyxl.Workbook()
    build_summary(wb, forecasts)
    build_problems(wb, forecasts)
    build_stats(wb, forecasts)

    wb.save(output_path)
    resolved = Path(output_path).resolve()
    logger.info("Saved → %s", resolved)
    print(f"\nOutput: {resolved}")


if __name__ == "__main__":
    main()
